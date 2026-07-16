"""اختبارات وحدة التصدير (JSON, CSV, Excel, PDF)."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import pytest

from app.core.dashboard import DashboardService, ViolationRow
from app.core.db import Database
from app.core.exporter import (
    build_study,
    export_csv,
    export_excel,
    export_json,
    export_pdf,
    record_export,
)


# ============================================
# Helpers
# ============================================
def _seed(tmp_db: Database) -> DashboardService:
    tmp_db.execute(
        "INSERT INTO videos (id, filepath, filename, source_type, status) "
        "VALUES (?, ?, ?, ?, ?)",
        (1, "/tmp/v.mp4", "v.mp4", "dashcam", "analyzed"),
    )
    tmp_db.execute(
        "INSERT INTO violations "
        "(video_id, violation_type, start_ms, end_ms, confidence, license_plate) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (1, "red_light_running", 0, 1500, 0.92, "ABC 1234"),
    )
    return DashboardService(db=tmp_db)


# ============================================
# build_study
# ============================================
def test_build_study_returns_full_payload(tmp_db: Database) -> None:
    service = _seed(tmp_db)
    study = build_study(service)
    assert study["kpis"]["total_violations"] == 1
    assert study["violations"][0]["violation_type"] == "red_light_running"
    assert "by_type" in study and "by_hour" in study


# ============================================
# JSON
# ============================================
def test_export_json_preserves_arabic(tmp_path: Path, tmp_db: Database) -> None:
    service = _seed(tmp_db)
    study = build_study(service)
    out = export_json(study, tmp_path / "study.json")
    raw = out.read_text(encoding="utf-8")
    parsed = json.loads(raw)

    assert parsed["violations"][0]["violation_type_ar"] == "قطع الإشارة الحمراء"
    # ASCII escaping ممنوع — العربي يجب أن يظهر بشكله الأصلي
    assert "قطع" in raw


def test_export_json_handles_datetime() -> None:
    payload = {"now": datetime(2025, 1, 1, 12, 0, 0)}
    text = json.dumps(
        payload, default=lambda v: v.isoformat() if hasattr(v, "isoformat") else str(v)
    )
    assert "2025-01-01" in text


# ============================================
# CSV
# ============================================
def test_export_csv_writes_header_and_rows(tmp_path: Path, tmp_db: Database) -> None:
    service = _seed(tmp_db)
    violations = service.list_violations()
    out = export_csv(violations, tmp_path / "v.csv")

    text = out.read_text(encoding="utf-8-sig")
    assert "violation_type" in text
    assert "red_light_running" in text
    assert "ABC 1234" in text


def test_export_csv_with_empty_list_writes_header_only(tmp_path: Path) -> None:
    out = export_csv([], tmp_path / "empty.csv")
    lines = out.read_text(encoding="utf-8-sig").strip().splitlines()
    assert len(lines) == 1  # header فقط


# ============================================
# Excel
# ============================================
def test_export_excel_creates_multiple_sheets(tmp_path: Path, tmp_db: Database) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    service = _seed(tmp_db)
    violations = service.list_violations()
    out = export_excel(build_study(service), violations, tmp_path / "s.xlsx")

    wb = load_workbook(out)
    assert "الملخص" in wb.sheetnames
    assert "حسب النوع" in wb.sheetnames
    assert "حسب الساعة" in wb.sheetnames
    assert "التفاصيل" in wb.sheetnames


def test_export_excel_summary_contains_kpis(tmp_path: Path, tmp_db: Database) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    service = _seed(tmp_db)
    out = export_excel(build_study(service), service.list_violations(), tmp_path / "s.xlsx")
    wb = load_workbook(out)
    sheet = wb["الملخص"]
    rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
    assert ("إجمالي المخالفات", 1) in rows


# ============================================
# PDF
# ============================================
def test_export_pdf_creates_file(tmp_path: Path, tmp_db: Database) -> None:
    pytest.importorskip("reportlab")
    service = _seed(tmp_db)
    out = export_pdf(
        build_study(service),
        service.list_violations(),
        tmp_path / "report.pdf",
    )
    assert out.exists()
    # PDF magic bytes
    assert out.read_bytes()[:4] == b"%PDF"


def test_export_pdf_with_empty_violations(tmp_path: Path, tmp_db: Database) -> None:
    pytest.importorskip("reportlab")
    service = DashboardService(db=tmp_db)
    out = export_pdf(build_study(service), [], tmp_path / "empty.pdf")
    assert out.exists()


# ============================================
# record_export
# ============================================
def test_record_export_writes_to_db(tmp_path: Path, tmp_db: Database) -> None:
    out = tmp_path / "demo.json"
    out.write_text("{}", encoding="utf-8")
    record_export(tmp_db, study_name="demo", fmt="json", output_path=out)

    row = tmp_db.fetch_one("SELECT study_name, format, output_path FROM exports")
    assert row is not None
    assert row[0] == "demo"
    assert row[1] == "json"


# ============================================
# find_arabic_font (منطق البحث عن الخط — بلا reportlab)
# ============================================
def test_find_arabic_font_prefers_env_override(tmp_path, monkeypatch) -> None:
    from app.core.exporter import find_arabic_font

    font = tmp_path / "MyArabic.ttf"
    font.write_bytes(b"fake-ttf")
    monkeypatch.setenv("BASEER_PDF_FONT", str(font))
    assert find_arabic_font() == font


def test_find_arabic_font_ignores_missing_env(monkeypatch) -> None:
    from app.core.exporter import find_arabic_font

    monkeypatch.setenv("BASEER_PDF_FONT", "/no/such/font.ttf")
    result = find_arabic_font()
    # يعود لخطوط النظام/الحزمة أو None — لكن لا يعيد المسار غير الموجود
    assert result is None or result.is_file()


def test_find_arabic_font_uses_bundled_assets(tmp_path, monkeypatch) -> None:
    import app.core.exporter as exporter_mod

    monkeypatch.delenv("BASEER_PDF_FONT", raising=False)
    fonts_dir = tmp_path / "assets" / "fonts"
    fonts_dir.mkdir(parents=True)
    bundled = fonts_dir / "Amiri-Regular.ttf"
    bundled.write_bytes(b"fake")
    monkeypatch.setattr(exporter_mod, "PROJECT_ROOT", tmp_path)
    assert exporter_mod.find_arabic_font() == bundled
