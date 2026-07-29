"""تصدير الدراسات بصيغ متعددة (JSON, CSV, Excel, PDF عربي)."""

from __future__ import annotations

import csv
import hashlib
import json
import logging
import os
from dataclasses import asdict, replace
from datetime import datetime
from pathlib import Path
from typing import Any

from app.config import PROJECT_ROOT
from app.constants import SOURCE_ARABIC_NAMES, VIOLATION_ARABIC_NAMES, SourceType, ViolationType
from app.core.dashboard import DashboardService, ViolationRow

logger = logging.getLogger(__name__)

# مسارات مرشّحة لخط عربي يدعم التشكيل — أول موجود يُستخدم في PDF
_SYSTEM_ARABIC_FONT_CANDIDATES: tuple[str, ...] = (
    "/usr/share/fonts/truetype/amiri/Amiri-Regular.ttf",
    "/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf",
    "/usr/share/fonts/truetype/kacst/KacstOne.ttf",
    "/Library/Fonts/Arial.ttf",
    "C:/Windows/Fonts/arial.ttf",
    "C:/Windows/Fonts/tahoma.ttf",
)


def find_arabic_font() -> Path | None:
    """يبحث عن ملف خط عربي (TTF) بترتيب: متغيّر بيئة → assets/fonts → خطوط النظام.

    يُرجع أول مسار موجود، أو None (فيعود PDF إلى Helvetica الذي لا يرسم العربية).
    دالة نقية قابلة للاختبار دون reportlab.
    """
    env = os.environ.get("BASEER_PDF_FONT")
    if env and Path(env).is_file():
        return Path(env)

    try:
        from app.config import get_settings

        configured = get_settings().pdf_font
        if configured and Path(configured).is_file():
            return Path(configured)
    except Exception:  # noqa: BLE001
        pass

    assets_fonts = PROJECT_ROOT / "assets" / "fonts"
    if assets_fonts.is_dir():
        for ttf in sorted(assets_fonts.glob("*.ttf")):
            return ttf

    for candidate in _SYSTEM_ARABIC_FONT_CANDIDATES:
        if Path(candidate).is_file():
            return Path(candidate)
    return None


def _register_arabic_font() -> str | None:
    """يسجّل الخط العربي في reportlab ويُرجع اسمه، أو None لو غير متاح."""
    font_path = find_arabic_font()
    if font_path is None:
        logger.warning(
            "لم يُعثر على خط عربي لـ PDF — النص العربي قد لا يظهر. "
            "ضع خطاً في assets/fonts/ أو عيّن BASEER_PDF_FONT."
        )
        return None
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        font_name = "BaseerArabic"
        pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
        logger.info("سُجّل خط PDF العربي: %s", font_path)
        return font_name
    except Exception as exc:  # noqa: BLE001
        logger.warning("تعذّر تسجيل خط PDF العربي %s: %s", font_path, exc)
        return None


# ============================================
# التجهيل (Anonymization)
# ============================================
# ملح ثابت للتجهيل داخل الدراسة الواحدة: يجعل نفس اللوحة تُعطي نفس الرمز
# (فتبقى التحليلات «كم مخالفة لنفس المركبة؟» ممكنة) دون كشف الرقم نفسه.
_DEFAULT_SALT_ENV = "BASEER_ANON_SALT"


def anonymization_salt() -> str:
    """الملح المستخدم في تجهيل اللوحات — من البيئة ثم من الإعدادات."""
    env_value = os.environ.get(_DEFAULT_SALT_ENV)
    if env_value:
        return env_value
    try:
        from app.config import get_settings

        return get_settings().anon_salt
    except Exception:  # noqa: BLE001 - لا نُسقط التصدير على إعدادات معطوبة
        return "baseer-default-salt"


def pseudonymize_plate(plate: str | None, *, salt: str | None = None) -> str | None:
    """يحوّل رقم لوحة إلى رمز مستعار ثابت لا يمكن عكسه.

    الدراسة الإحصائية (توزيع المخالفات على الأنواع والساعات) **لا تحتاج أرقام
    اللوحات أصلاً**، بينما تصديرها يجعل ملف CSV سجلاً شخصياً كاملاً يربط مركبات
    محدَّدة بأوقات ومواقع. الرمز المستعار يحفظ قابلية التجميع ويُسقط الهوية.
    """
    if not plate:
        return None
    digest = hashlib.sha256(f"{salt or anonymization_salt()}::{plate}".encode()).hexdigest()
    return f"PLATE-{digest[:10].upper()}"


def anonymize_violation_rows(
    violations: list[ViolationRow], *, salt: str | None = None
) -> list[ViolationRow]:
    """ينسخ صفوف المخالفات بلوحات مستعارة (بلا تعديل الأصل)."""
    return [
        replace(v, license_plate=pseudonymize_plate(v.license_plate, salt=salt)) for v in violations
    ]


def anonymize_study(study: dict[str, Any], *, salt: str | None = None) -> dict[str, Any]:
    """ينسخ الدراسة مع تجهيل اللوحات داخل قائمة المخالفات."""
    out: dict[str, Any] = dict(study)
    out["anonymized"] = True
    out["violations"] = [
        {**v, "license_plate": pseudonymize_plate(v.get("license_plate"), salt=salt)}
        for v in study.get("violations", [])
    ]
    return out


# ============================================
# تجميع دراسة كاملة
# ============================================
def build_study(service: DashboardService, *, anonymize: bool = False) -> dict[str, Any]:
    """يجمع كل بيانات الدراسة في dict واحد قابل للتسلسل.

    `anonymize=True` يستبدل أرقام اللوحات برموز مستعارة ثابتة — للنشر والمشاركة.
    """
    kpis = service.get_kpis()
    study = {
        "generated_at": datetime.now().isoformat(),
        "kpis": asdict(kpis),
        "violations": [asdict(v) for v in service.list_violations()],
        "by_type": service.violations_by_type(),
        "by_hour": service.violations_by_hour(),
        "by_weekday": service.violations_by_weekday(),
        "by_review_status": service.violations_by_review_status(),
    }
    return anonymize_study(study) if anonymize else study


# ============================================
# JSON
# ============================================
def export_json(study: dict[str, Any], output_path: Path | str) -> Path:
    """يكتب الدراسة كـ JSON بدعم عربي كامل."""
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(
        json.dumps(study, ensure_ascii=False, indent=2, default=_json_default),
        encoding="utf-8",
    )
    return out


def _json_default(value: object) -> str:
    if isinstance(value, datetime):
        return value.isoformat()
    isoformat = getattr(value, "isoformat", None)
    if callable(isoformat):
        return str(isoformat())
    return str(value)


# ============================================
# CSV
# ============================================
CSV_COLUMNS: tuple[str, ...] = (
    "id",
    "video_id",
    "video_filename",
    "violation_type",
    "violation_type_ar",
    "start_ms",
    "end_ms",
    "confidence",
    "license_plate",
    "review_status",
    "notes",
)


def export_csv(violations: list[ViolationRow], output_path: Path | str) -> Path:
    """يصدّر قائمة المخالفات كـ CSV."""
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for v in violations:
            row = {col: getattr(v, col, "") for col in CSV_COLUMNS}
            writer.writerow(row)
    return out


# ============================================
# Excel (لو openpyxl متاح)
# ============================================
def export_excel(
    study: dict[str, Any],
    violations: list[ViolationRow],
    output_path: Path | str,
) -> Path:
    """يصدّر الدراسة كـ Excel متعدد الـ sheets."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("يحتاج openpyxl — ثبّت requirements.txt") from exc

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Summary sheet
    summary = wb.active
    summary.title = "الملخص"
    kpis = study.get("kpis", {})
    summary.append(["مؤشر", "القيمة"])
    summary.append(["إجمالي المقاطع", kpis.get("total_videos", 0)])
    summary.append(["إجمالي المخالفات", kpis.get("total_violations", 0)])
    summary.append(["متوسط مخالفات/مقطع", round(kpis.get("avg_violations_per_video", 0.0), 2)])
    summary.append([])
    summary.append(["المصدر", "عدد المقاطع"])
    for src, count in (kpis.get("sources_breakdown") or {}).items():
        ar = _source_arabic_name(src)
        summary.append([ar, count])

    # By type sheet
    by_type_sheet = wb.create_sheet("حسب النوع")
    by_type_sheet.append(["النوع (الكود)", "النوع (عربي)", "العدد"])
    for vtype, count in study.get("by_type", []):
        by_type_sheet.append([vtype, _violation_arabic_name(vtype), count])

    # By hour sheet
    by_hour_sheet = wb.create_sheet("حسب الساعة")
    by_hour_sheet.append(["الساعة", "عدد المخالفات"])
    for hour, count in study.get("by_hour", []):
        by_hour_sheet.append([hour, count])

    # Details sheet
    details = wb.create_sheet("التفاصيل")
    details.append(list(CSV_COLUMNS))
    for v in violations:
        details.append([getattr(v, col, "") for col in CSV_COLUMNS])

    wb.save(out)
    return out


# ============================================
# PDF عربي (reportlab + arabic-reshaper + bidi)
# ============================================
def export_pdf(
    study: dict[str, Any],
    violations: list[ViolationRow],
    output_path: Path | str,
    *,
    title: str = "تقرير بَصير — تحليل المخالفات المرورية",
) -> Path:
    """يُصدّر تقرير PDF عربي RTL مع الرسوم الأساسية كجداول."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer
    except ImportError as exc:
        raise RuntimeError("يحتاج reportlab — ثبّت requirements.txt") from exc

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(out),
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        title="Baseer Study",
    )

    arabic_font = _register_arabic_font()
    styles = getSampleStyleSheet()
    body_font = arabic_font or styles["Normal"].fontName
    rtl_style = ParagraphStyle(
        "rtl",
        parent=styles["Normal"],
        alignment=2,  # right
        fontName=body_font,
        fontSize=11,
        leading=16,
    )
    title_style = ParagraphStyle(
        "rtl_title",
        parent=styles["Heading1"],
        alignment=2,
        fontName=body_font,
        fontSize=18,
        leading=22,
    )

    story: list[Any] = [Paragraph(_shape(title), title_style), Spacer(1, 12)]

    kpis = study.get("kpis", {})
    story.append(Paragraph(_shape("المؤشرات الرئيسية"), rtl_style))
    story.append(_kpi_table(kpis, arabic_font))
    story.append(Spacer(1, 12))

    story.append(Paragraph(_shape("المخالفات حسب النوع"), rtl_style))
    story.append(_by_type_table(study.get("by_type", []), arabic_font))
    story.append(Spacer(1, 12))

    if violations:
        story.append(Paragraph(_shape("أحدث المخالفات (آخر 30)"), rtl_style))
        story.append(_violations_table(violations[:30], arabic_font))

    doc.build(story)
    return out


def _kpi_table(kpis: dict[str, Any], font: str | None = None) -> object:
    from reportlab.lib import colors
    from reportlab.platypus import Table

    data = [
        [_shape("المؤشر"), _shape("القيمة")],
        [_shape("إجمالي المقاطع"), str(kpis.get("total_videos", 0))],
        [_shape("إجمالي المخالفات"), str(kpis.get("total_violations", 0))],
        [_shape("متوسط مخالفات/مقطع"), f"{kpis.get('avg_violations_per_video', 0.0):.2f}"],
    ]
    table = Table(data, colWidths=[8, 4], hAlign="RIGHT")
    table.setStyle(_default_table_style(colors, font))
    return table


def _by_type_table(by_type: list[Any], font: str | None = None) -> object:
    from reportlab.lib import colors
    from reportlab.platypus import Table

    data = [[_shape("النوع"), _shape("العدد")]]
    for vtype, count in by_type:
        data.append([_shape(_violation_arabic_name(vtype)), str(count)])
    table = Table(data, colWidths=[10, 3], hAlign="RIGHT")
    table.setStyle(_default_table_style(colors, font))
    return table


def _violations_table(violations: list[ViolationRow], font: str | None = None) -> object:
    from reportlab.lib import colors
    from reportlab.platypus import Table

    data = [
        [
            _shape("الملف"),
            _shape("النوع"),
            _shape("الثقة"),
            _shape("الحالة"),
        ]
    ]
    for v in violations:
        data.append(
            [
                _shape(v.video_filename),
                _shape(v.violation_type_ar),
                f"{v.confidence:.2f}",
                _shape(v.review_status),
            ]
        )
    table = Table(data, colWidths=[6, 4, 2, 3], hAlign="RIGHT")
    table.setStyle(_default_table_style(colors, font))
    return table


def _default_table_style(colors_module: Any, font: str | None = None) -> object:
    from reportlab.platypus import TableStyle

    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors_module.HexColor("#2c3e50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors_module.white),
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors_module.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
    ]
    if font is not None:
        commands.append(("FONTNAME", (0, 0), (-1, -1), font))
    return TableStyle(commands)


def _shape(text: str) -> str:
    """يُعيد تشكيل النص العربي للعرض الصحيح في PDF."""
    if not text:
        return ""
    try:
        from app.utils.arabic_utils import shape_for_pdf

        return shape_for_pdf(text)
    except Exception:  # noqa: BLE001
        return text


def _violation_arabic_name(vtype: str) -> str:
    try:
        return VIOLATION_ARABIC_NAMES[ViolationType(vtype)]
    except (KeyError, ValueError):
        return vtype


def _source_arabic_name(source: str) -> str:
    try:
        return SOURCE_ARABIC_NAMES[SourceType(source)]
    except (KeyError, ValueError):
        return source


# ============================================
# تسجيل التصدير في DB
# ============================================
def record_export(
    db: Any,
    *,
    study_name: str,
    fmt: str,
    output_path: Path,
    filter_json: str | None = None,
) -> None:
    """يُسجّل تصديراً في جدول exports."""
    db.execute(
        "INSERT INTO exports (study_name, filter_json, format, output_path) " "VALUES (?, ?, ?, ?)",
        (study_name, filter_json, fmt, str(output_path)),
    )


__all__ = [
    "anonymization_salt",
    "anonymize_study",
    "anonymize_violation_rows",
    "build_study",
    "export_csv",
    "export_excel",
    "export_json",
    "export_pdf",
    "pseudonymize_plate",
    "record_export",
]
